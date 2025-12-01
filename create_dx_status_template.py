#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DX化対応状況管理テンプレート.xlsx を作成するスクリプト
マトリクス比較表と進捗ダッシュボードを含むエクセルファイルを生成します
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_dx_status_template():
    """DX化対応状況管理テンプレートを作成"""
    wb = Workbook()
    
    # シート1: マトリクス比較表
    ws_matrix = wb.active
    ws_matrix.title = "マトリクス比較表"
    
    # シート2: 進捗ダッシュボード
    ws_dashboard = wb.create_sheet("進捗ダッシュボード")
    
    # シート3: Before/After比較
    ws_before_after = wb.create_sheet("Before_After比較")
    
    # スタイル定義
    title_font = Font(name="メイリオ", size=16, bold=True, color="FFFFFF")
    heading_font = Font(name="メイリオ", size=12, bold=True)
    body_font = Font(name="メイリオ", size=10)
    
    title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    heading_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    complete_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    in_progress_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    not_started_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========================================
    # シート1: マトリクス比較表
    # ========================================
    
    # タイトル
    ws_matrix.merge_cells("A1:H1")
    title_cell = ws_matrix["A1"]
    title_cell.value = "DX化対応状況 - マトリクス比較表"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_alignment
    ws_matrix.row_dimensions[1].height = 30
    
    # 更新日
    ws_matrix["A2"] = f"更新日: {datetime.now().strftime('%Y年%m月%d日')}"
    ws_matrix["A2"].font = Font(name="メイリオ", size=9)
    
    # ヘッダー行
    headers = [
        "業務プロセス", "現状の方法", "キントーン対応方法", 
        "対応状況", "優先度", "担当者", "完了予定", "備考"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_matrix.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = heading_font
        cell.fill = heading_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    ws_matrix.row_dimensions[3].height = 25
    
    # サンプルデータ
    sample_data = [
        ["生産日報管理", "Excel手入力", "生産日報報告書アプリ", "✅ 完了", "高", "田中", "2025/01", "自動集計機能追加"],
        ["不良管理", "紙媒体", "不良名マスタアプリ", "🔄 進行中", "高", "佐藤", "2025/02", "データ移行中"],
        ["在庫管理", "別システム", "在庫管理アプリ", "⏳ 未着手", "中", "鈴木", "2025/03", "要件定義中"],
        ["品質検査", "手作業記録", "品質検査アプリ", "⏳ 未着手", "中", "山田", "2025/04", "設計中"],
        ["設備管理", "紙ベース", "設備管理アプリ", "⏳ 未着手", "低", "高橋", "2025/05", "検討中"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_matrix.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = body_font
            cell.alignment = left_alignment if col_idx in [1, 2, 3, 8] else center_alignment
            cell.border = thin_border
            
            # 対応状況に応じて背景色を設定
            if col_idx == 4:  # 対応状況列
                if "✅" in str(value):
                    cell.fill = complete_fill
                elif "🔄" in str(value):
                    cell.fill = in_progress_fill
                elif "⏳" in str(value):
                    cell.fill = not_started_fill
        
        ws_matrix.row_dimensions[row_idx].height = 30
    
    # 列幅の調整
    column_widths = {
        "A": 20,  # 業務プロセス
        "B": 25,  # 現状の方法
        "C": 25,  # キントーン対応方法
        "D": 15,  # 対応状況
        "E": 10,  # 優先度
        "F": 12,  # 担当者
        "G": 15,  # 完了予定
        "H": 30,  # 備考
    }
    
    for col, width in column_widths.items():
        ws_matrix.column_dimensions[col].width = width
    
    # 凡例
    legend_row = len(sample_data) + 5
    ws_matrix.merge_cells(f"A{legend_row}:H{legend_row}")
    legend_cell = ws_matrix[f"A{legend_row}"]
    legend_cell.value = "【凡例】✅ 完了 | 🔄 進行中 | ⏳ 未着手 | ⚠️ 課題あり | ❌ 保留"
    legend_cell.font = Font(name="メイリオ", size=9, italic=True)
    legend_cell.alignment = left_alignment
    
    # ========================================
    # シート2: 進捗ダッシュボード
    # ========================================
    
    # タイトル
    ws_dashboard.merge_cells("A1:D1")
    title_cell = ws_dashboard["A1"]
    title_cell.value = "DX化進捗状況ダッシュボード"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_alignment
    ws_dashboard.row_dimensions[1].height = 30
    
    # 更新日
    ws_dashboard["A2"] = f"更新日: {datetime.now().strftime('%Y年%m月%d日')}"
    ws_dashboard["A2"].font = Font(name="メイリオ", size=9)
    
    # 全体進捗
    row = 4
    ws_dashboard.merge_cells(f"A{row}:D{row}")
    progress_title = ws_dashboard[f"A{row}"]
    progress_title.value = "【全体進捗】"
    progress_title.font = heading_font
    progress_title.fill = heading_fill
    progress_title.alignment = left_alignment
    ws_dashboard.row_dimensions[row].height = 25
    
    row += 1
    ws_dashboard["A5"] = "完了"
    ws_dashboard["B5"] = "2件"
    ws_dashboard["B5"].fill = complete_fill
    ws_dashboard["C5"] = "40%"
    
    row += 1
    ws_dashboard["A6"] = "進行中"
    ws_dashboard["B6"] = "1件"
    ws_dashboard["B6"].fill = in_progress_fill
    ws_dashboard["C6"] = "20%"
    
    row += 1
    ws_dashboard["A7"] = "未着手"
    ws_dashboard["B7"] = "2件"
    ws_dashboard["B7"].fill = not_started_fill
    ws_dashboard["C7"] = "40%"
    
    # カテゴリ別進捗
    row += 2
    ws_dashboard.merge_cells(f"A{row}:D{row}")
    category_title = ws_dashboard[f"A{row}"]
    category_title.value = "【カテゴリ別進捗】"
    category_title.font = heading_font
    category_title.fill = heading_fill
    category_title.alignment = left_alignment
    ws_dashboard.row_dimensions[row].height = 25
    
    categories = [
        ["生産管理", "2/3", "67%"],
        ["品質管理", "1/2", "50%"],
        ["在庫管理", "0/1", "0%"],
    ]
    
    for idx, (category, progress, percentage) in enumerate(categories, start=row+1):
        ws_dashboard[f"A{idx}"] = category
        ws_dashboard[f"B{idx}"] = progress
        ws_dashboard[f"C{idx}"] = percentage
        ws_dashboard[f"A{idx}"].font = body_font
        ws_dashboard[f"B{idx}"].font = body_font
        ws_dashboard[f"C{idx}"].font = body_font
    
    # 列幅の調整
    ws_dashboard.column_dimensions["A"].width = 20
    ws_dashboard.column_dimensions["B"].width = 15
    ws_dashboard.column_dimensions["C"].width = 15
    ws_dashboard.column_dimensions["D"].width = 15
    
    # ========================================
    # シート3: Before/After比較
    # ========================================
    
    # タイトル
    ws_before_after.merge_cells("A1:D1")
    title_cell = ws_before_after["A1"]
    title_cell.value = "Before/After比較表"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_alignment
    ws_before_after.row_dimensions[1].height = 30
    
    # ヘッダー行
    before_after_headers = ["業務フロー", "Before（現状）", "After（キントーン）", "改善効果"]
    
    for col_idx, header in enumerate(before_after_headers, start=1):
        cell = ws_before_after.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = heading_font
        cell.fill = heading_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    ws_before_after.row_dimensions[3].height = 25
    
    # サンプルデータ
    before_after_data = [
        ["データ入力", "Excel手入力 → メール送信", "キントーンアプリで直接入力", "時間短縮: 50%"],
        ["承認フロー", "紙で回覧 → 押印", "ワークフロー機能で自動承認", "処理時間: 70%短縮"],
        ["データ集計", "手動でExcel集計", "自動集計・グラフ表示", "ミス削減: 90%"],
        ["レポート作成", "毎月手作業で作成", "ダッシュボードで自動生成", "作業時間: 80%削減"],
    ]
    
    for row_idx, row_data in enumerate(before_after_data, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_before_after.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = body_font
            cell.alignment = left_alignment
            cell.border = thin_border
        
        ws_before_after.row_dimensions[row_idx].height = 40
    
    # 列幅の調整
    ws_before_after.column_dimensions["A"].width = 20
    ws_before_after.column_dimensions["B"].width = 35
    ws_before_after.column_dimensions["C"].width = 35
    ws_before_after.column_dimensions["D"].width = 25
    
    # ファイルを保存
    output_file = "DX化対応状況管理テンプレート.xlsx"
    wb.save(output_file)
    print(f"DX化対応状況管理テンプレートを作成しました: {output_file}")
    print("\n含まれるシート:")
    print("1. マトリクス比較表 - 詳細な対応状況を管理")
    print("2. 進捗ダッシュボード - 全体の進捗を可視化")
    print("3. Before_After比較 - 改善効果を明確化")

if __name__ == "__main__":
    try:
        create_dx_status_template()
    except ImportError:
        print("エラー: openpyxlライブラリが必要です。")
        print("インストール方法: pip install openpyxl")
    except Exception as e:
        print(f"エラーが発生しました: {e}")

