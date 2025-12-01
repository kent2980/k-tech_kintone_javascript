#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
キントーン_データ登録手順書.xlsx を作成するスクリプト
画像を貼り付けて手順を説明するエクセルファイルを生成します
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

def create_manual():
    """キントーンへのデータ登録手順書を作成"""
    wb = Workbook()
    ws = wb.active
    ws.title = "データ登録手順"
    
    # スタイル定義
    title_font = Font(name="メイリオ", size=16, bold=True, color="FFFFFF")
    heading_font = Font(name="メイリオ", size=14, bold=True)
    body_font = Font(name="メイリオ", size=11)
    
    title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    heading_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # タイトル
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "キントーン 不良名マスタ データ登録手順書"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_alignment
    ws.row_dimensions[1].height = 30
    
    # 手順1: エクセルでデータを準備
    row = 3
    ws.merge_cells(f"A{row}:D{row}")
    step1_cell = ws[f"A{row}"]
    step1_cell.value = "手順1: エクセルでデータを準備"
    step1_cell.font = heading_font
    step1_cell.fill = heading_fill
    step1_cell.alignment = left_alignment
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    desc1_cell = ws[f"A{row}"]
    desc1_cell.value = "エクセルファイルに「ライン名」と「不良名」の2列を作成し、データを入力します。\nデータは末尾の空行に追加してください。"
    desc1_cell.font = body_font
    desc1_cell.alignment = left_alignment
    ws.row_dimensions[row].height = 50
    
    # 画像1の説明（エクセルデータの例）
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    img1_desc = ws[f"A{row}"]
    img1_desc.value = "【画像1】エクセルでのデータ準備例\n・列A: ライン名（例: A38, A35, A33, A06, A01など）\n・列B: 不良名（例: Time NG, 逆, シルク欠けなど）\n・データは末尾の空行（57行目以降）に追加"
    img1_desc.font = body_font
    img1_desc.alignment = left_alignment
    ws.row_dimensions[row].height = 80
    
    # 画像1のプレースホルダー
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    img1_placeholder = ws[f"A{row}"]
    img1_placeholder.value = "【ここに画像1を貼り付け】\nエクセルスプレッドシートの画面（44-56行目にデータ、57-59行目が空行）"
    img1_placeholder.font = Font(name="メイリオ", size=10, italic=True, color="808080")
    img1_placeholder.alignment = center_alignment
    img1_placeholder.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws.row_dimensions[row].height = 200
    
    # 手順2: キントーンアプリを開く
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    step2_cell = ws[f"A{row}"]
    step2_cell.value = "手順2: キントーンの不良名マスタアプリを開く"
    step2_cell.font = heading_font
    step2_cell.fill = heading_fill
    step2_cell.alignment = left_alignment
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    desc2_cell = ws[f"A{row}"]
    desc2_cell.value = "キントーンの「不良名マスタ」アプリを開きます。\nアプリが空の状態（データが表示されていない状態）から開始します。"
    desc2_cell.font = body_font
    desc2_cell.alignment = left_alignment
    ws.row_dimensions[row].height = 50
    
    # 画像2の説明
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    img2_desc = ws[f"A{row}"]
    img2_desc.value = "【画像2】不良名マスタアプリ（データなし）\n・アプリタイトル: 不良名マスタ\n・テーブルに「データがありません。」と表示されている状態"
    img2_desc.font = body_font
    img2_desc.alignment = left_alignment
    ws.row_dimensions[row].height = 60
    
    # 画像2のプレースホルダー
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    img2_placeholder = ws[f"A{row}"]
    img2_placeholder.value = "【ここに画像2を貼り付け】\nキントーン不良名マスタアプリの画面（データなし）"
    img2_placeholder.font = Font(name="メイリオ", size=10, italic=True, color="808080")
    img2_placeholder.alignment = center_alignment
    img2_placeholder.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws.row_dimensions[row].height = 200
    
    # 手順3: ファイルから読み込むを選択
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    step3_cell = ws[f"A{row}"]
    step3_cell.value = "手順3: 「ファイルから読み込む」を選択"
    step3_cell.font = heading_font
    step3_cell.fill = heading_fill
    step3_cell.alignment = left_alignment
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    desc3_cell = ws[f"A{row}"]
    desc3_cell.value = "画面右上の「...」（三点リーダー）ボタンをクリックし、\nメニューから「ファイルから読み込む」を選択します。"
    desc3_cell.font = body_font
    desc3_cell.alignment = left_alignment
    ws.row_dimensions[row].height = 50
    
    # 画像3の説明
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    img3_desc = ws[f"A{row}"]
    img3_desc.value = "【画像3】メニューから「ファイルから読み込む」を選択\n・右上の「...」ボタンをクリック\n・メニューが開いたら「ファイルから読み込む」を選択"
    img3_desc.font = body_font
    img3_desc.alignment = left_alignment
    ws.row_dimensions[row].height = 60
    
    # 画像3のプレースホルダー
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    img3_placeholder = ws[f"A{row}"]
    img3_placeholder.value = "【ここに画像3を貼り付け】\nメニューが開いた状態（「ファイルから読み込む」が選択可能）"
    img3_placeholder.font = Font(name="メイリオ", size=10, italic=True, color="808080")
    img3_placeholder.alignment = center_alignment
    img3_placeholder.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws.row_dimensions[row].height = 200
    
    # 手順4: ファイルを選択
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    step4_cell = ws[f"A{row}"]
    step4_cell.value = "手順4: エクセルファイルを選択"
    step4_cell.font = heading_font
    step4_cell.fill = heading_fill
    step4_cell.alignment = left_alignment
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    desc4_cell = ws[f"A{row}"]
    desc4_cell.value = "ファイル選択ダイアログが開いたら、\n手順1で作成したエクセルファイル（.xlsxまたは.csv）を選択します。"
    desc4_cell.font = body_font
    desc4_cell.alignment = left_alignment
    ws.row_dimensions[row].height = 50
    
    # 画像4の説明
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    img4_desc = ws[f"A{row}"]
    img4_desc.value = "【画像4】ファイル選択ダイアログ\n・ファイル一覧からエクセルファイルを選択\n・対応形式: .xlsx, .csv, .txt"
    img4_desc.font = body_font
    img4_desc.alignment = left_alignment
    ws.row_dimensions[row].height = 60
    
    # 画像4のプレースホルダー
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    img4_placeholder = ws[f"A{row}"]
    img4_placeholder.value = "【ここに画像4を貼り付け】\nファイル選択ダイアログの画面"
    img4_placeholder.font = Font(name="メイリオ", size=10, italic=True, color="808080")
    img4_placeholder.alignment = center_alignment
    img4_placeholder.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws.row_dimensions[row].height = 200
    
    # 手順5: 読み込むボタンをクリック
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    step5_cell = ws[f"A{row}"]
    step5_cell.value = "手順5: 「読み込む」ボタンをクリック"
    step5_cell.font = heading_font
    step5_cell.fill = heading_fill
    step5_cell.alignment = left_alignment
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    desc5_cell = ws[f"A{row}"]
    desc5_cell.value = "ファイルを選択したら、「読み込む」ボタンをクリックします。\nデータがキントーンに登録されます。"
    desc5_cell.font = body_font
    desc5_cell.alignment = left_alignment
    ws.row_dimensions[row].height = 50
    
    # 画像5の説明
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    img5_desc = ws[f"A{row}"]
    img5_desc.value = "【画像5】読み込み画面\n・「キャンセル」ボタン: 読み込みを中止\n・「読み込む」ボタン: データを読み込んで登録"
    img5_desc.font = body_font
    img5_desc.alignment = left_alignment
    ws.row_dimensions[row].height = 60
    
    # 画像5のプレースホルダー
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    img5_placeholder = ws[f"A{row}"]
    img5_placeholder.value = "【ここに画像5を貼り付け】\n読み込み画面（キャンセルと読み込むボタン）"
    img5_placeholder.font = Font(name="メイリオ", size=10, italic=True, color="808080")
    img5_placeholder.alignment = center_alignment
    img5_placeholder.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws.row_dimensions[row].height = 200
    
    # 完了
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    complete_cell = ws[f"A{row}"]
    complete_cell.value = "完了: データが正常に登録されると、不良名マスタアプリにデータが表示されます。"
    complete_cell.font = Font(name="メイリオ", size=12, bold=True, color="006100")
    complete_cell.alignment = center_alignment
    complete_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws.row_dimensions[row].height = 40
    
    # 列幅の調整
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    
    # ファイルを保存
    output_file = "キントーン_データ登録手順書.xlsx"
    wb.save(output_file)
    print(f"手順書を作成しました: {output_file}")
    print("\n注意: 画像は手動で貼り付けてください。")
    print("各手順の【ここに画像Xを貼り付け】のセルに、対応する画像を貼り付けてください。")

if __name__ == "__main__":
    try:
        create_manual()
    except ImportError:
        print("エラー: openpyxlライブラリが必要です。")
        print("インストール方法: pip install openpyxl")
    except Exception as e:
        print(f"エラーが発生しました: {e}")

